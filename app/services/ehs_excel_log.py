"""EHS master log — one _MasterLog.xlsx per form type in OneDrive.

Ported from metfraa-ehs excel-log.js: on every approval/rejection we download
the log (if it exists), append a row, and re-upload. Same columns, same
brand styling, so existing logs keep growing seamlessly.
"""
import io
import logging
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from . import onedrive

log = logging.getLogger(__name__)

BRAND_BLUE = "005B96"
GREEN = "1F8B4C"
RED = "C0392B"

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def ehs_root() -> str:
    return os.getenv("EHS_ONEDRIVE_ROOT", "Metfraa-EHS").strip("/")


def _headers_for(form: dict) -> list[str]:
    headers = ["Submission ID", "Submitted At", "Submitted By (Name)", "Submitted By (Email)"]
    for f in form["fields"]:
        headers.append(f"{f['label']} (link)" if f["type"] == "photo" else f["label"])
    for i, item in enumerate(form.get("checklist") or []):
        headers.append(f"#{i + 1} {item} — Result")
        headers.append(f"#{i + 1} Remarks")
        headers.append(f"#{i + 1} Photo (link)")
    headers += ["PDF Report (link)", "Status", "Reviewed By (Name)", "Reviewed By (Email)",
                "Reviewed At", "Edits Made", "Reject Reason"]
    return headers


def append_to_master_log(form: dict, sub, file_links: dict, pdf_link: str | None) -> None:
    """sub: EHSSubmission ORM object. file_links: {"fields": {key: [url]}, "checklist": {idx: [url]}}."""
    log_path = f"{ehs_root()}/{form['folder']}/_MasterLog.xlsx"

    existing = onedrive.download_from_path(log_path)
    if existing:
        wb = load_workbook(io.BytesIO(existing))
        ws = wb["Submissions"] if "Submissions" in wb.sheetnames else wb.create_sheet("Submissions")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Submissions"
        hdr = _headers_for(form)
        ws.append(hdr)
        fill = PatternFill("solid", fgColor=BRAND_BLUE)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF", size=11)
            c.fill = fill
            c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        ws.row_dimensions[1].height = 32
        for idx, h in enumerate(hdr, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max(len(h) + 2, 14), 60)

    fields = sub.fields or {}
    checklist = sub.checklist or []
    fl = (file_links or {}).get("fields", {})
    cl = (file_links or {}).get("checklist", {})

    row: list = [sub.submission_id, sub.submitted_at_ist, sub.submitted_by_name, sub.submitted_by_email or ""]
    for f in form["fields"]:
        if f["type"] == "photo":
            row.append(", ".join(fl.get(f["key"], [])) or "")
        else:
            v = fields.get(f["key"], "")
            row.append(", ".join(v) if isinstance(v, list) else (v if v is not None else ""))
    for i in range(len(form.get("checklist") or [])):
        item = checklist[i] if i < len(checklist) else {}
        row.append((item or {}).get("result", ""))
        row.append((item or {}).get("remarks", ""))
        links = cl.get(i, cl.get(str(i), []))
        row.append(", ".join(links) if links else "")
    row.append(pdf_link or "")
    row.append("Approved" if sub.status == "approved" else "Rejected")
    row.append(sub.reviewed_by_name or "")
    row.append(sub.reviewed_by_email or "")
    row.append(sub.reviewed_at_ist or "")
    row.append(sub.edits_made or "")
    row.append(sub.reject_reason or "")

    ws.append(row)
    status_cell = ws.cell(row=ws.max_row, column=len(row) - 5)
    status_cell.font = Font(bold=True, color="FFFFFF")
    status_cell.fill = PatternFill("solid", fgColor=GREEN if sub.status == "approved" else RED)

    buf = io.BytesIO()
    wb.save(buf)
    onedrive.upload_to_path(buf.getvalue(), log_path, XLSX_CT)
    log.info(f"[ehs-log] appended {sub.submission_id} to {log_path}")
