"""Builds the master Excel file containing all employee daily entries."""
import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from ..models import Employee, KPI, DailyEntry, KPIEntry


METFRAA_BLUE = "3B82F6"
METFRAA_BLACK = "0A0A0A"


def _header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=METFRAA_BLACK)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="888888"),
        right=Side(style="thin", color="888888"),
        top=Side(style="thin", color="888888"),
        bottom=Side(style="thin", color="888888"),
    )


def _data_style(cell, alt: bool = False):
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    if alt:
        cell.fill = PatternFill("solid", fgColor="F3F4F6")


def build_master_workbook(db: Session) -> bytes:
    """Generate the full master Excel workbook in-memory.

    Layout:
    - Sheet 1: Summary (employee × month aggregate)
    - Sheet per employee: daily entries with each KPI as a column
    """
    wb = Workbook()

    # Remove default sheet — we'll add our own
    default = wb.active
    wb.remove(default)

    # ---- Summary sheet ----
    ws = wb.create_sheet("Summary")
    headers = [
        "Employee", "Designation", "Email",
        "Total Work Days", "Leave Days", "Site/Remote", "Sundays", "Holidays",
        "Total Entries",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        _header_style(cell)
    ws.row_dimensions[1].height = 28

    employees = db.query(Employee).filter_by(is_active=True).order_by(Employee.name).all()
    for r_idx, emp in enumerate(employees, start=2):
        entries = db.query(DailyEntry).filter_by(employee_id=emp.id).all()
        counts = {"work": 0, "casual_leave": 0, "site_remote": 0, "sunday": 0, "holiday": 0}
        for e in entries:
            counts[e.entry_type] = counts.get(e.entry_type, 0) + 1
        row = [
            emp.name, emp.designation, emp.email,
            counts.get("work", 0),
            counts.get("casual_leave", 0),
            counts.get("site_remote", 0),
            counts.get("sunday", 0),
            counts.get("holiday", 0),
            len(entries),
        ]
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            _data_style(cell, alt=(r_idx % 2 == 0))

    # Column widths
    widths = [25, 32, 30, 14, 12, 14, 10, 10, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # ---- Per-employee sheet ----
    for emp in employees:
        sheet_name = (emp.name[:28] + "..") if len(emp.name) > 30 else emp.name
        ws = wb.create_sheet(sheet_name)

        kpis = (
            db.query(KPI)
            .filter_by(employee_id=emp.id, is_active=True)
            .order_by(KPI.display_order)
            .all()
        )
        kpi_headers = [f"{k.name} ({k.unit})" for k in kpis]
        headers = ["Date", "Type", "Comments"] + kpi_headers
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            _header_style(cell)
        ws.row_dimensions[1].height = 36

        entries = (
            db.query(DailyEntry)
            .filter_by(employee_id=emp.id)
            .order_by(DailyEntry.entry_date)
            .all()
        )
        for r_idx, entry in enumerate(entries, start=2):
            row = [
                entry.entry_date.isoformat(),
                entry.entry_type.replace("_", " ").title(),
                entry.comments or "",
            ]
            # Fetch KPI values for this day
            kpi_value_map = {kv.kpi_id: kv.value for kv in entry.kpi_values}
            for k in kpis:
                row.append(kpi_value_map.get(k.id, 0) if entry.entry_type == "work" else "")

            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                _data_style(cell, alt=(r_idx % 2 == 0))

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 40
        for i in range(4, 4 + len(kpis)):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 22

        ws.freeze_panes = "D2"

    # Output to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
