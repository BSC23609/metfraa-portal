"""Expense OneDrive artifacts — approval PDF + per-form master log.

Root folder defaults to "Reimbursements and Conveyance" (the existing expense
folder in info@metfraa.com's drive); override with EXPENSE_ONEDRIVE_ROOT.

Layout:
  <root>/<YYYY-MM>/<reference>/Bills/…           (bills, uploaded at submit)
  <root>/<YYYY-MM>/<reference>/<reference>.pdf   (approval report)
  <root>/_MasterLog_<FORM>.xlsx                  (one log per form type)
"""
import io
import logging
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from . import onedrive

log = logging.getLogger(__name__)

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
BRAND = colors.HexColor("#005B96")
LIGHT = colors.HexColor("#eef2f6")

S_TITLE = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=15, textColor=colors.white)
S_SUB = ParagraphStyle("s", fontName="Helvetica", fontSize=9, textColor=colors.HexColor("#cfe3f2"))
S_SEC = ParagraphStyle("sec", fontName="Helvetica-Bold", fontSize=11, textColor=BRAND, spaceBefore=8, spaceAfter=4)
S_L = ParagraphStyle("l", fontName="Helvetica-Bold", fontSize=8.5, textColor=colors.HexColor("#6b7480"))
S_V = ParagraphStyle("v", fontName="Helvetica", fontSize=9.5, textColor=colors.HexColor("#1a2332"))
S_C = ParagraphStyle("c", fontName="Helvetica", fontSize=8.5, textColor=colors.HexColor("#1a2332"))
S_SM = ParagraphStyle("sm", fontName="Helvetica", fontSize=8, textColor=colors.HexColor("#6b7480"))


def expense_root() -> str:
    return os.getenv("EXPENSE_ONEDRIVE_ROOT", "Reimbursements and Conveyance").strip("/")


def submission_folder(sub) -> str:
    return f"{expense_root()}/{sub.period or 'no-period'}/{sub.reference}"


# ------------------------------------------------------------------ PDF

def _kv(label, value, width):
    return Table([[Paragraph(label, S_L), Paragraph(str(value) if value not in (None, "") else "—", S_V)]],
                 colWidths=[width * 0.32, width * 0.68],
                 style=TableStyle([("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#dde3ea")),
                                   ("TOPPADDING", (0, 0), (-1, -1), 3),
                                   ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))


def _rows_table(headers, rows, width, weights):
    data = [[Paragraph(f"<b>{h}</b>", S_C) for h in headers]]
    for r in rows:
        data.append([Paragraph(str(c) if c not in (None, "") else "—", S_C) for c in r])
    return Table(data, colWidths=[width * w for w in weights], repeatRows=1,
                 style=TableStyle([
                     ("BACKGROUND", (0, 0), (-1, 0), BRAND),
                     ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                     ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c9d3dd")),
                     ("VALIGN", (0, 0), (-1, -1), "TOP"),
                     ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
                     ("TOPPADDING", (0, 0), (-1, -1), 3),
                     ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                 ]))


def _payload_elements(form_type: str, p: dict, width) -> list:
    E = []
    inr = "₹{:,.2f}".format
    if form_type == "met_local":
        E.append(_kv("Vehicle", f"{p.get('vehicle_label')} @ ₹{p.get('rate_per_km')}/km  {p.get('vehicle_reg', '')}", width))
        E.append(Spacer(1, 4))
        E.append(_rows_table(["Date", "From", "To", "Purpose", "KM", "Amount"],
                             [[t["date"], t["from"], t["to"], t.get("purpose", ""), f"{t['km']:g}", inr(t["amount"])] for t in p.get("trips", [])],
                             width, (0.13, 0.2, 0.2, 0.25, 0.08, 0.14)))
    elif form_type == "met_cab":
        E.append(_rows_table(["Date", "Pickup", "Drop", "KM", "Fare", "Purpose"],
                             [[r["date"], r["pickup"], r["drop"], f"{r['km']:g}", inr(r["fare"]), r.get("purpose", "")] for r in p.get("rides", [])],
                             width, (0.12, 0.2, 0.2, 0.08, 0.14, 0.26)))
    elif form_type == "met_accommodation":
        E.append(_kv("Level / Daily limit", f"{p.get('level')} — ₹{p.get('daily_limit')}/day", width))
        E.append(Spacer(1, 4))
        E.append(_rows_table(["Date", "Location", "Hotel", "Bill No.", "Amount"],
                             [[e["date"], e["location"], e.get("hotel", ""), e.get("bill_no", ""), inr(e["amount"])] for e in p.get("entries", [])],
                             width, (0.14, 0.24, 0.26, 0.16, 0.2)))
    elif form_type == "met_outstation":
        for t in p.get("trips", []):
            E.append(_kv("Trip", f"{t['place']}  ({t['from_date']} → {t['to_date']}) — {t['purpose']}", width))
            rows = []
            for cat, items in (t.get("categories") or {}).items():
                for it in items:
                    rows.append([it["date"], cat.replace("_", " ").title(), it.get("desc", ""), inr(it["amount"])])
            if rows:
                E.append(Spacer(1, 3))
                E.append(_rows_table(["Date", "Category", "Description", "Amount"], rows,
                                     width, (0.14, 0.2, 0.44, 0.22)))
            E.append(Spacer(1, 5))
    elif form_type == "met_misc":
        E.append(_rows_table(["Date", "Purpose", "Amount"],
                             [[i["date"], i["purpose"], inr(i["amount"])] for i in p.get("items", [])],
                             width, (0.16, 0.6, 0.24)))
    elif form_type == "met_advance":
        for k, label in [("destination", "Destination"), ("travel_from", "Travel From"),
                         ("travel_to", "Travel To"), ("purpose", "Purpose"),
                         ("mode", "Mode"), ("notes", "Notes"), ("amount", "Advance Amount (₹)")]:
            E.append(_kv(label, p.get(k), width))
    elif form_type == "met_dtr":
        E.append(_rows_table(["Date", "Mode", "From", "To", "Purpose", "Fare"],
                             [[e["date"], e["mode"].replace("_", " ").title(), e["from"], e["to"],
                               (e.get("client_name") or e.get("purpose_other_reason") or e["purpose_category"].replace("_", " ").title()),
                               inr(e["fare"])] for e in p.get("entries", [])],
                             width, (0.12, 0.12, 0.18, 0.18, 0.26, 0.14)))
    return E


def generate_expense_pdf(sub, form_title: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=12 * mm, bottomMargin=14 * mm, title=sub.reference)
    W = doc.width
    E = [Table([[Paragraph("METFRAA — Expense Report", S_TITLE)],
                [Paragraph(f"{form_title} · {sub.reference}", S_SUB)]],
               colWidths=[W],
               style=TableStyle([("BACKGROUND", (0, 0), (-1, -1), BRAND),
                                 ("LEFTPADDING", (0, 0), (-1, -1), 10),
                                 ("TOPPADDING", (0, 0), (0, 0), 8),
                                 ("BOTTOMPADDING", (0, 1), (0, 1), 8)])),
         Spacer(1, 6),
         _kv("Employee", f"{sub.employee_name}  ({sub.employee_email or ''})", W),
         _kv("Period", sub.period or "—", W),
         _kv("Submitted (IST)", sub.submitted_at_ist, W),
         _kv("Total Claimed", "₹{:,.2f}".format(sub.total_amount), W),
         Spacer(1, 4),
         Paragraph("Details", S_SEC)]
    E += _payload_elements(sub.form_type, sub.payload or {}, W)
    E.append(Spacer(1, 8))
    status_label = "APPROVED" if sub.status in ("approved", "advance_approved", "settled") else sub.status.upper()
    E.append(Table([[Paragraph(status_label, ParagraphStyle("ap", fontName="Helvetica-Bold", fontSize=11, textColor=colors.white)),
                     Paragraph(f"Reviewed by <b>{sub.reviewed_by or ''}</b> at {sub.reviewed_at_ist or ''} IST"
                               + (f"<br/>Note: {sub.review_note}" if sub.review_note else ""), S_V)]],
                   colWidths=[W * 0.2, W * 0.8],
                   style=TableStyle([("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#1F8B4C")),
                                     ("BACKGROUND", (1, 0), (1, -1), LIGHT),
                                     ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                     ("TOPPADDING", (0, 0), (-1, -1), 6),
                                     ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                                     ("LEFTPADDING", (0, 0), (-1, -1), 8)])))
    E.append(Spacer(1, 6))
    E.append(Paragraph(f"Generated by Metfraa Portal · {sub.reference}", S_SM))
    doc.build(E)
    return buf.getvalue()


# ------------------------------------------------------------------ master log

LOG_HEADERS = ["Reference", "Employee", "Email", "Level", "Period", "Submitted At",
               "Total (₹)", "Status", "Reviewed By", "Reviewed At", "Note / Changes Required",
               "Bills", "PDF Report (link)"]


def append_expense_log(sub, form_code: str, bill_links: list[str], pdf_link: str | None) -> None:
    path = f"{expense_root()}/_MasterLog_{form_code}.xlsx"
    existing = onedrive.download_from_path(path)
    if existing:
        wb = load_workbook(io.BytesIO(existing))
        ws = wb["Submissions"] if "Submissions" in wb.sheetnames else wb.create_sheet("Submissions")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Submissions"
        ws.append(LOG_HEADERS)
        fill = PatternFill("solid", fgColor="005B96")
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF", size=11)
            c.fill = fill
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30
        for idx, h in enumerate(LOG_HEADERS, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max(len(h) + 4, 14), 50)
    ws.append([sub.reference, sub.employee_name, sub.employee_email or "", sub.employee_level or "",
               sub.period or "", sub.submitted_at_ist, sub.total_amount,
               sub.status, sub.reviewed_by or "", sub.reviewed_at_ist or "",
               sub.review_note or sub.changes_required or "",
               ", ".join(bill_links) if bill_links else "", pdf_link or ""])
    status_cell = ws.cell(row=ws.max_row, column=8)
    approvedish = sub.status in ("approved", "advance_approved", "settled")
    status_cell.font = Font(bold=True, color="FFFFFF")
    status_cell.fill = PatternFill("solid", fgColor="1F8B4C" if approvedish else ("C0392B" if sub.status == "rejected" else "B7791F"))
    buf = io.BytesIO()
    wb.save(buf)
    onedrive.upload_to_path(buf.getvalue(), path, XLSX_CT)
    log.info(f"[expense-log] appended {sub.reference} to {path}")


# ============================================================================
# Consolidated monthly report PDF
#
# Deliberate deviation from the source, flagged in the port spec: the Node
# builder MERGED each claim's stored PDF and its bill files into one document
# using pdf-lib. Doing that here would mean downloading N PDFs from OneDrive
# plus every bill, then merging — on Vercel that is both slow (60s ceiling)
# and fragile (one unreachable bill breaks the whole report).
#
# Instead the report is RENDERED in one pass from the submission rows: cover,
# contents table, then a detail block per claim. Same information, one build,
# no network fan-out. Bills stay linked from each claim's own PDF in OneDrive.
# ============================================================================

def _inr(n) -> str:
    try:
        v = float(n or 0)
    except (TypeError, ValueError):
        v = 0.0
    return f"Rs. {v:,.2f}"


def build_consolidated_pdf(report, employee, rows: list) -> tuple[bytes, int]:
    """rows: the ExpenseSubmission objects included in this report.

    Returns (pdf_bytes, page_count).
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)

    INK = colors.HexColor("#1a2332")
    MUTED = colors.HexColor("#6b7280")
    BRAND = colors.HexColor("#005B96")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm,
                            topMargin=18 * mm, bottomMargin=16 * mm,
                            title=f"Consolidated Report {report.period}")
    h1 = ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=20, textColor=INK,
                        spaceAfter=2)
    lbl = ParagraphStyle("lbl", fontName="Helvetica-Bold", fontSize=8, textColor=MUTED,
                         spaceAfter=2)
    big = ParagraphStyle("big", fontName="Helvetica-Bold", fontSize=15, textColor=INK)
    body = ParagraphStyle("body", fontName="Helvetica", fontSize=9.5, textColor=INK,
                          leading=13)
    sec = ParagraphStyle("sec", fontName="Helvetica-Bold", fontSize=12, textColor=BRAND,
                         spaceBefore=6, spaceAfter=4)

    total = float(report.total_amount or 0)
    negative = total < 0
    el = [
        Paragraph("CONSOLIDATED REPORT", h1),
        Paragraph(report.period or "", ParagraphStyle("p", fontName="Helvetica",
                                                      fontSize=12, textColor=MUTED)),
        Spacer(1, 8 * mm),
        Paragraph("EMPLOYEE", lbl),
        Paragraph(employee.get("name") or "-", big),
        Paragraph(" · ".join(x for x in (employee.get("email"), employee.get("code")) if x),
                  body),
        Spacer(1, 6 * mm),
        Paragraph("NET OWED BACK BY EMPLOYEE" if negative else "TOTAL PAYABLE", lbl),
        Paragraph(_inr(abs(total)), ParagraphStyle(
            "tot", fontName="Helvetica-Bold", fontSize=22,
            textColor=colors.HexColor("#b91c1c") if negative else INK)),
        Paragraph(f"{report.submission_count} claim"
                  f"{'' if report.submission_count == 1 else 's'}", body),
        Spacer(1, 6 * mm),
        Paragraph("STATUS", lbl),
        Paragraph((report.status or "").replace("_", " ").title(), big),
        Spacer(1, 6 * mm),
    ]
    for label, who, when in (("Sent by HR", report.hr_approved_by, report.hr_approved_at),
                             ("Approved by Mgmt", report.mgmt_approved_by,
                              report.mgmt_approved_at)):
        if who:
            el.append(Paragraph(f"<b>{label}:</b> {who} · {when or ''}", body))
    el.append(Spacer(1, 8 * mm))

    # ---- contents ----
    el.append(Paragraph("Contents", sec))
    head = ["#", "Reference", "Form", "Submitted", "Status", "Amount"]
    data = [head]
    for i, s in enumerate(rows, 1):
        if s.status == "settled" and s.form_type == "met_advance":
            amt = float(s.differential_amount or 0)
        elif s.status == "settled":
            amt = float((s.actuals or {}).get("actual_amount") or s.total_amount or 0)
        else:
            amt = float(s.total_amount or 0)
        data.append([str(i), s.reference, (s.form_type or "").replace("met_", ""),
                     (s.submitted_at_ist or "")[:10],
                     (s.status or "").replace("_", " "), _inr(amt)])
    t = Table(data, colWidths=[10 * mm, 42 * mm, 26 * mm, 24 * mm, 30 * mm, 30 * mm],
              repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (5, 0), (5, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f8fa")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    el.append(t)

    # ---- per-claim detail ----
    for i, s in enumerate(rows, 1):
        el.append(PageBreak())
        el.append(Paragraph(f"{i}. {s.reference}", sec))
        meta = [["Form", (s.form_type or "").replace("met_", "")],
                ["Period", s.period or "-"],
                ["Submitted", s.submitted_at_ist or "-"],
                ["Status", (s.status or "").replace("_", " ")],
                ["Claimed", _inr(s.total_amount)]]
        if s.status == "settled":
            meta.append(["Actuals", _inr((s.actuals or {}).get("actual_amount"))])
            if s.differential_amount is not None:
                meta.append(["Differential", _inr(s.differential_amount)])
        if s.reviewed_by:
            meta.append(["Reviewed by", s.reviewed_by])
        mt = Table(meta, colWidths=[32 * mm, 130 * mm])
        mt.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e5e7eb")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        el.append(mt)
        el.append(Spacer(1, 4 * mm))
        try:
            el.extend(_payload_elements(s.form_type, s.payload or {}, 162 * mm))
        except Exception:
            el.append(Paragraph("(claim detail unavailable)", body))
        if s.pdf_web_url:
            el.append(Spacer(1, 3 * mm))
            el.append(Paragraph(
                f'<font color="#005B96"><a href="{s.pdf_web_url}">'
                "Open the individual claim PDF (with bills) in OneDrive</a></font>", body))

    doc.build(el)
    return buf.getvalue(), doc.page


def _safe_name(name: str) -> str:
    """Match the OLD Node app's safeName() exactly, so a legacy OneDrive path
    reconstructs byte-for-byte: forbidden chars -> '-', whitespace collapsed,
    trimmed, capped at 120."""
    import re
    out = re.sub(r'[\\/:*?"<>|]', "-", str(name or "Unknown"))
    out = re.sub(r"\s+", " ", out).strip()
    return out[:120]


def legacy_employee_folder(name: str, code: str | None) -> str:
    """<root>/<name> (<code>) — the layout the Node app synced to, and where
    every MIGRATED claim's files actually live. New portal claims use
    submission_folder() instead; this exists only to reach the old files."""
    label = _safe_name(name)
    if code:
        label += f" ({_safe_name(code)})"
    return f"{expense_root()}/{label}"


def legacy_report_path(name: str, code: str | None, reference: str) -> str:
    return f"{legacy_employee_folder(name, code)}/Reports/{reference}.pdf"


def legacy_bill_path(name: str, code: str | None, reference: str, filename: str) -> str:
    return f"{legacy_employee_folder(name, code)}/Uploads/{reference}__{filename}"
