"""Daily consolidated task report — builds an Excel + emails it to VP.

Called by scheduler at 10:00 AM IST daily (skips Sundays) and available
on-demand via admin endpoint for any past date.

Excel layout (one row per task, color-coded per employee):
  Employee | Code | Department | Task # | Task Description | Status | Project | Remarks | Tomorrow's Plan | Blockers
"""
import io
import logging
import os
from datetime import date, datetime, timedelta

import pytz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..database import SessionLocal
from ..models import DailyTaskReport, Employee
from .email_service import send_email_async
from .onedrive import upload_file

log = logging.getLogger(__name__)

IST = pytz.timezone(os.getenv("TIMEZONE", "Asia/Kolkata"))

# Recipients
TO_EMAIL = os.getenv("DAILY_REPORT_TO", "vp@metfraa.com")
CC_EMAILS = [e.strip() for e in os.getenv(
    "DAILY_REPORT_CC", "admin@metfraa.com,arasu@metfraa.com"
).split(",") if e.strip()]

# OneDrive folder for archives
ARCHIVE_FOLDER_TEMPLATE = "DailyReports/{year_month}"

# ----------------------------------------------------------------
# Color palette — one stable color per employee code (index into palette)
# Soft pastels that print well and are easy on the eye
# ----------------------------------------------------------------
PALETTE = [
    "FDE7E7", "FDF2E7", "FDF9E7", "F0FDE7", "E7FDF0", "E7FDF9",
    "E7F5FD", "E7EBFD", "F0E7FD", "F9E7FD", "FDE7F5", "FDE7EB",
    "FFF3E7", "FFF8E7", "F4FFE7", "E7FFE7", "E7FFF3", "E7FFFA",
    "E7F7FF", "E7EEFF", "F3E7FF", "FAE7FF", "FFE7F7", "FFE7EE",
    "FDF0E7", "F7FDE7", "E7FDF7", "E7F0FD", "F0E7FA", "FDE7F0",
]

MISSED_FILL = "FFD6D6"  # muted red
HEADER_FILL = "1F2937"  # slate-900
HEADER_TEXT = "FFFFFF"


def _color_for(code: str) -> str:
    """Deterministic palette color for an employee code."""
    if not code:
        return "FFFFFF"
    h = sum(ord(c) for c in code)
    return PALETTE[h % len(PALETTE)]


def _thin_border() -> Border:
    side = Side(style="thin", color="D0D5DD")
    return Border(top=side, bottom=side, left=side, right=side)


# ----------------------------------------------------------------
# Data gathering
# ----------------------------------------------------------------
def _gather_data(db: Session, target: date) -> dict:
    """Fetch all employees + their reports for a date.

    Returns:
      {
        "employees": [Employee ordered by department, name],
        "reports_by_emp": {emp_id: DailyTaskReport or None},
      }
    """
    employees = (
        db.query(Employee)
        .filter(Employee.is_active.is_(True))
        .filter(Employee.can_submit_task_report.is_(True))
        .order_by(Employee.department.asc().nulls_last(), Employee.name.asc())
        .all()
    )
    reports = (
        db.query(DailyTaskReport)
        .filter(DailyTaskReport.report_date == target)
        .all()
    )
    reports_by_emp = {r.employee_id: r for r in reports}
    return {"employees": employees, "reports_by_emp": reports_by_emp}


# ----------------------------------------------------------------
# Excel builder
# ----------------------------------------------------------------
def build_excel(target: date, db: Session) -> tuple[bytes, dict]:
    """Build the consolidated Excel workbook. Returns (bytes, stats_dict)."""
    data = _gather_data(db, target)
    employees = data["employees"]
    reports_by_emp = data["reports_by_emp"]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{target.strftime('%d-%b-%Y')}"

    # ---- Title row ----
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"Daily Task Report — {target.strftime('%A, %d %B %Y')}"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # ---- Column headers ----
    headers = [
        "Employee", "Code", "Department", "Task #",
        "Task Description", "Status", "Project", "Remarks",
        "Tomorrow's Plan", "Blockers",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color=HEADER_TEXT, size=11)
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin_border()
    ws.row_dimensions[2].height = 32

    # ---- Column widths ----
    widths = [24, 10, 18, 8, 42, 12, 20, 30, 30, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Data rows ----
    row = 3
    submitters = 0
    total_tasks = 0
    total_completed = 0
    total_pending = 0
    missed = 0
    empty_reports = 0

    for emp in employees:
        report = reports_by_emp.get(emp.id)
        emp_color = _color_for(emp.employee_code or emp.name)
        emp_fill = PatternFill("solid", fgColor=emp_color)

        if report is None:
            # Missed submission → one red row
            missed += 1
            _write_row(
                ws, row,
                emp_name=emp.name,
                code=emp.employee_code or "",
                dept=emp.department or "",
                task_num="",
                task_desc="NOT SUBMITTED",
                status="",
                project="",
                remarks="",
                tomorrow="",
                blockers="",
                fill=PatternFill("solid", fgColor=MISSED_FILL),
                bold=True,
                italic=True,
            )
            row += 1
            continue

        submitters += 1
        items = sorted(report.items, key=lambda i: i.sequence)
        if not items:
            # Empty submission (only plan/blockers) → one row
            empty_reports += 1
            _write_row(
                ws, row,
                emp_name=emp.name,
                code=emp.employee_code or "",
                dept=emp.department or "",
                task_num="",
                task_desc="(no tasks entered)",
                status="",
                project="",
                remarks="",
                tomorrow=report.tomorrow_plan or "",
                blockers=report.blockers or "",
                fill=emp_fill,
                italic=True,
            )
            row += 1
            continue

        # Normal — one row per task
        for idx, item in enumerate(items, start=1):
            total_tasks += 1
            if item.status == "completed":
                total_completed += 1
            else:
                total_pending += 1
            # Only show name/code/dept/plan/blockers on the FIRST row per employee
            first = idx == 1
            _write_row(
                ws, row,
                emp_name=emp.name if first else "",
                code=(emp.employee_code or "") if first else "",
                dept=(emp.department or "") if first else "",
                task_num=idx,
                task_desc=item.task_description,
                status=item.status,
                project=item.project or "",
                remarks=item.remarks or "",
                tomorrow=(report.tomorrow_plan or "") if first else "",
                blockers=(report.blockers or "") if first else "",
                fill=emp_fill,
            )
            row += 1

    # ---- Footer / summary ----
    total_emps = len(employees)
    summary_row = row + 1
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=10)
    summary_cell = ws.cell(row=summary_row, column=1)
    summary_cell.value = (
        f"Summary — {submitters}/{total_emps} employees submitted "
        f"· {total_tasks} tasks ({total_completed} completed, {total_pending} pending) "
        f"· {missed} missed submissions"
        + (f" · {empty_reports} submitted without tasks" if empty_reports else "")
    )
    summary_cell.font = Font(italic=True, size=10, color="555555")
    summary_cell.alignment = Alignment(horizontal="left")

    ws.freeze_panes = "A3"

    # Serialize
    buf = io.BytesIO()
    wb.save(buf)

    stats = {
        "total_employees": total_emps,
        "submitters": submitters,
        "missed": missed,
        "empty_reports": empty_reports,
        "total_tasks": total_tasks,
        "total_completed": total_completed,
        "total_pending": total_pending,
    }
    return buf.getvalue(), stats


def _write_row(ws, row, *, emp_name, code, dept, task_num, task_desc, status,
               project, remarks, tomorrow, blockers, fill=None, bold=False, italic=False):
    """Write one row with the standard styling."""
    values = [
        emp_name, code, dept, task_num, task_desc, status,
        project, remarks, tomorrow, blockers,
    ]
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = _thin_border()
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.font = Font(
            size=10,
            bold=bold or (col == 1),  # employee name always bold
            italic=italic,
            color="B91C1C" if (fill and fill.fgColor.rgb == MISSED_FILL) else "111827",
        )
        if fill:
            cell.fill = fill

    # Status column color coding on the value itself
    if status:
        status_cell = ws.cell(row=row, column=6)
        status_cell.font = Font(
            size=10,
            bold=True,
            color="065F46" if status == "completed" else "92400E",
        )
        status_cell.value = "✓ Completed" if status == "completed" else "○ Pending"


# ----------------------------------------------------------------
# Email HTML body
# ----------------------------------------------------------------
def _email_html(target: date, stats: dict, onedrive_url: str | None) -> str:
    date_str = target.strftime("%A, %d %B %Y")
    return f"""
    <html><body style="font-family:-apple-system,Segoe UI,Roboto,sans-serif;background:#f3f4f6;padding:24px">
      <div style="max-width:640px;margin:auto;background:white;border-radius:8px;overflow:hidden;border:1px solid #e5e7eb">
        <div style="background:#0a0a0a;padding:18px 24px;border-bottom:3px solid #1E3A8A">
          <div style="color:white;font-size:18px;font-weight:bold;letter-spacing:1px">METFRAA</div>
          <div style="color:#9ca3af;font-size:11px">Steeling the Future</div>
        </div>
        <div style="padding:28px 24px">
          <div style="font-size:12px;color:#6b7280;text-transform:uppercase;letter-spacing:0.5px;font-weight:700">Daily Task Report</div>
          <h2 style="margin:6px 0 12px;color:#0a0a0a;font-size:22px">{date_str}</h2>

          <table style="width:100%;border-collapse:separate;border-spacing:0;margin-top:18px">
            <tr>
              <td style="background:#f0fdf4;border:1px solid #86efac;border-radius:6px;padding:12px 14px;width:33%">
                <div style="font-size:11px;color:#166534;text-transform:uppercase;letter-spacing:0.5px;font-weight:700">Submitted</div>
                <div style="font-size:22px;font-weight:700;color:#166534;margin-top:4px">{stats['submitters']}/{stats['total_employees']}</div>
              </td>
              <td style="width:8px"></td>
              <td style="background:#eff6ff;border:1px solid #93c5fd;border-radius:6px;padding:12px 14px;width:33%">
                <div style="font-size:11px;color:#1e40af;text-transform:uppercase;letter-spacing:0.5px;font-weight:700">Total tasks</div>
                <div style="font-size:22px;font-weight:700;color:#1e40af;margin-top:4px">{stats['total_tasks']}</div>
                <div style="font-size:11px;color:#3b82f6;margin-top:2px">{stats['total_completed']} completed · {stats['total_pending']} pending</div>
              </td>
              <td style="width:8px"></td>
              <td style="background:#fef2f2;border:1px solid #fca5a5;border-radius:6px;padding:12px 14px;width:33%">
                <div style="font-size:11px;color:#991b1b;text-transform:uppercase;letter-spacing:0.5px;font-weight:700">Not submitted</div>
                <div style="font-size:22px;font-weight:700;color:#991b1b;margin-top:4px">{stats['missed']}</div>
              </td>
            </tr>
          </table>

          <p style="color:#374151;line-height:1.5;margin-top:22px;font-size:14px">
            The consolidated task report for {date_str} is attached as an Excel file with one row per task,
            color-coded per employee for easier reading.
          </p>

          {'<p style="color:#374151;font-size:12px;margin-top:12px">' + f'<a href="{onedrive_url}" style="color:#1E3A8A;text-decoration:none;font-weight:600">Open in OneDrive →</a>' + '</p>' if onedrive_url else ''}
        </div>
        <div style="background:#f9fafb;padding:12px 24px;color:#6b7280;font-size:11px;text-align:center">
          Automated report from Metfraa KPI Tracker · Sent at 10:00 AM IST
        </div>
      </div>
    </body></html>
    """


# ----------------------------------------------------------------
# Main entry — used by scheduler and admin endpoint
# ----------------------------------------------------------------
async def generate_and_dispatch(
    target: date,
    *,
    upload_to_onedrive: bool = True,
    send_email: bool = True,
) -> dict:
    """Build the daily Excel for `target` date, upload to OneDrive, email VP.

    Returns a stats dict describing what happened. Safe to call from cron
    or from an admin-triggered endpoint.
    """
    db = SessionLocal()
    try:
        log.info(f"[daily-excel] building for {target.isoformat()}")
        xlsx_bytes, stats = build_excel(target, db)
        filename = f"MetfraaTasks_{target.strftime('%Y-%m-%d')}.xlsx"
        log.info(f"[daily-excel] built {filename}, {len(xlsx_bytes):,} bytes")

        # Upload to OneDrive under KPI_Tracker/DailyReports/YYYY-MM/
        onedrive_url = None
        if upload_to_onedrive:
            year_month = target.strftime("%Y-%m")
            folder_path = f"KPI_Tracker/DailyReports/{year_month}"
            try:
                result = upload_file(
                    xlsx_bytes,
                    filename,
                    folder_path,
                )
                if result and isinstance(result, dict):
                    onedrive_url = result.get("webUrl") or result.get("web_url")
                log.info(f"[daily-excel] uploaded to OneDrive: {folder_path}/{filename}")
            except Exception as e:
                log.error(f"[daily-excel] OneDrive upload failed: {e}")

        # Send email with attachment
        email_ok = False
        if send_email:
            subject = f"Daily Task Report — {target.strftime('%A, %d %b %Y')}"
            html = _email_html(target, stats, onedrive_url)
            try:
                email_ok = await send_email_async(
                    to=TO_EMAIL,
                    subject=subject,
                    html_body=html,
                    cc=CC_EMAILS,
                    attachments=[(filename, xlsx_bytes,
                                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")],
                )
                log.info(f"[daily-excel] email sent: {email_ok}")
            except Exception as e:
                log.error(f"[daily-excel] email failed: {e}")

        return {
            "target": target.isoformat(),
            "stats": stats,
            "filename": filename,
            "excel_bytes": len(xlsx_bytes),
            "onedrive_url": onedrive_url,
            "email_sent": email_ok,
        }
    finally:
        db.close()


def daily_task_report_job():
    """Scheduler entry point — build + dispatch yesterday's report.

    Skips Sundays (report_date being Sunday means no working day expected).
    """
    yesterday = datetime.now(IST).date() - timedelta(days=1)
    if yesterday.weekday() == 6:  # Sunday
        log.info(f"[daily-excel-cron] {yesterday} was a Sunday — skipping.")
        return

    import asyncio
    try:
        loop = asyncio.new_event_loop()
        try:
            result = loop.run_until_complete(generate_and_dispatch(yesterday))
            log.info(f"[daily-excel-cron] done: {result}")
        finally:
            loop.close()
    except Exception as e:
        log.error(f"[daily-excel-cron] failed: {e}", exc_info=True)
