"""Phase 3 data migration — run once against the live Neon DB.

Usage (from repo root, with DATABASE_URL pointing at Neon):

  1) Expense history from the live bsg-portal SQLite:
       python scripts/migrate_phase3.py expense /path/to/bsg-portal.db [--dry-run]

  2) EHS history from the Metfraa-EHS OneDrive master logs
     (needs MS_CLIENT_ID/SECRET/TENANT + ONEDRIVE_USER_EMAIL env vars):
       python scripts/migrate_phase3.py ehs [--dry-run]

Both are IDEMPOTENT — re-running skips anything already imported
(matched on reference / submission_id), so partial runs are safe.

Mapping notes:
- Employees matched by email (case-insensitive), then by employee_code.
  Unmatched submitters are reported and their rows imported with
  submitted_by_id=NULL so nothing is lost.
- bsg-portal levels JUNIOR/SENIOR/MANAGER -> L1/L2/L3 written to
  expense_employee_meta (only when no level is set yet).
- Only metfraa-company rows are imported (bsc_* forms skipped).
- EHS master-log rows import as approved/rejected submissions with
  fields parsed back from the log columns; photo/PDF OneDrive links
  preserved in the photos index / pdf_web_url.
"""
import argparse
import io
import json
import os
import sqlite3
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.database import SessionLocal  # noqa: E402
from app.models import (  # noqa: E402
    EHSSubmission, Employee, ExpenseAttachment, ExpenseEmployeeMeta,
    ExpenseMonthlyPayment, ExpenseProject, ExpenseSubmission,
)

LEVEL_MAP = {"JUNIOR": "L1", "SENIOR": "L2", "MANAGER": "L3",
             "L1": "L1", "L2": "L2", "L3": "L3"}


def _employee_maps(db):
    by_email, by_code = {}, {}
    for e in db.query(Employee).all():
        if e.email:
            by_email[e.email.strip().lower()] = e
        if e.employee_code:
            by_code[e.employee_code.strip().upper()] = e
    return by_email, by_code


# ================================================================ expense

def migrate_expense(sqlite_path: str, dry: bool) -> None:
    if not os.path.exists(sqlite_path):
        sys.exit(f"SQLite file not found: {sqlite_path}")
    src = sqlite3.connect(sqlite_path)
    src.row_factory = sqlite3.Row
    db = SessionLocal()
    by_email, by_code = _employee_maps(db)

    # ---- employees: map + set levels
    unmatched = []
    src_emps = {r["id"]: r for r in src.execute("SELECT * FROM employees WHERE company='metfraa'")}
    id_map = {}
    meta_done: set[int] = set()
    for sid, r in src_emps.items():
        e = by_email.get((r["email"] or "").strip().lower()) or by_code.get((r["employee_code"] or "").strip().upper())
        if e:
            id_map[sid] = e
            level = LEVEL_MAP.get((r["level"] or "").upper())
            if (level and e.id not in meta_done
                    and not db.query(ExpenseEmployeeMeta).filter_by(employee_id=e.id).first()):
                meta_done.add(e.id)
                if not dry:
                    db.add(ExpenseEmployeeMeta(employee_id=e.id, level=level))
        else:
            unmatched.append(f"{r['name']} <{r['email']}> ({r['employee_code']})")
    print(f"[expense] employees: {len(id_map)} matched, {len(unmatched)} unmatched")
    for u in unmatched:
        print(f"  ! no portal employee for: {u} — their rows import with no employee link")

    # ---- projects
    proj_map = {}
    for r in src.execute("SELECT * FROM projects"):
        p = db.query(ExpenseProject).filter(ExpenseProject.name == r["name"]).first()
        if not p and not dry:
            p = ExpenseProject(code=r["code"], name=r["name"], is_active=bool(r["is_active"]))
            db.add(p)
            db.flush()
        if p:
            proj_map[r["id"]] = p.id
    print(f"[expense] projects: {len(proj_map)} mapped")

    # ---- submissions
    existing = {ref for (ref,) in db.query(ExpenseSubmission.reference).all()}
    n_new = n_skip = 0
    att_rows = list(src.execute("SELECT * FROM attachments"))
    atts_by_sub = {}
    for a in att_rows:
        atts_by_sub.setdefault(a["submission_id"], []).append(a)
    pay_map = {}
    for r in src.execute("SELECT * FROM submissions WHERE company='metfraa'"):
        if r["reference"] in existing:
            n_skip += 1
            continue
        emp = id_map.get(r["employee_id"])
        try:
            payload = json.loads(r["payload_json"] or "{}")
        except json.JSONDecodeError:
            payload = {}
        # remap project ids inside DTR entries
        for e in (payload.get("entries") or []):
            if isinstance(e, dict) and e.get("project_id") in proj_map:
                e["project_id"] = proj_map[e["project_id"]]
        # Metfraa-expenses stores submission-level purpose fields as columns —
        # fold them into payload so nothing is lost (keys prefixed to avoid collisions)
        rk = r.keys()
        for col in ("purpose_category", "project_id", "client_name", "purpose_other_reason"):
            if col in rk and r[col] not in (None, "") and col not in payload:
                payload[col] = proj_map.get(r[col], r[col]) if col == "project_id" else r[col]
        try:
            actuals = json.loads(r["actuals_json"]) if r["actuals_json"] else None
        except (json.JSONDecodeError, TypeError):
            actuals = None
        sub = ExpenseSubmission(
            reference=r["reference"],
            employee_id=emp.id if emp else None,
            employee_name=(src_emps.get(r["employee_id"]) or {"name": "Unknown"})["name"],
            employee_email=(src_emps.get(r["employee_id"]) or {"email": None})["email"],
            employee_level=LEVEL_MAP.get(((src_emps.get(r["employee_id"]) or {"level": ""})["level"] or "").upper(), "L1"),
            form_type=r["form_type"], period=r["period"],
            payload=payload, total_amount=r["total_amount"] or 0,
            status=r["status"] or "pending",
            reviewed_by=r["reviewed_by"], reviewed_at_ist=r["reviewed_at"],
            review_note=r["review_note"], changes_required=r["changes_required"],
            returned_at_ist=r["returned_at"], actuals=actuals,
            settled_at_ist=r["settled_at"], settlement_reviewed_by=r["settlement_reviewed_by"],
            settlement_note=r["settlement_note"],
            pdf_web_url=None,  # old local pdf paths are not portable
            submitted_at_ist=r["submitted_at"],
        )
        if not dry:
            db.add(sub)
            db.flush()
            for a in atts_by_sub.get(r["id"], []):
                db.add(ExpenseAttachment(
                    submission_id=sub.id, filename=a["filename"],
                    onedrive_path=f"__legacy__/{a['stored_path']}",
                    web_url=None, mime_type=a["mime_type"],
                    size_bytes=a["size_bytes"], row_idx=a["row_idx"],
                    label=a["label"] or "legacy",
                ))
        if emp:
            pay_map.setdefault(r["employee_id"], emp)
        n_new += 1
    print(f"[expense] submissions: {n_new} imported, {n_skip} already present")
    print("  note: legacy bill files stay on the old server disk (paths kept with __legacy__/ prefix);"
          " copy them into OneDrive later if needed — new claims are unaffected")

    # ---- monthly payments
    n_pay = 0
    for r in src.execute("SELECT * FROM monthly_payments"):
        emp = id_map.get(r["employee_id"])
        if not emp:
            continue
        if db.query(ExpenseMonthlyPayment).filter_by(employee_id=emp.id, year=r["year"], month=r["month"]).first():
            continue
        if not dry:
            db.add(ExpenseMonthlyPayment(employee_id=emp.id, year=r["year"], month=r["month"],
                                         amount_paid=r["amount_paid"], paid_by=r["paid_by"],
                                         paid_at_ist=r["paid_at"]))
        n_pay += 1
    print(f"[expense] monthly payments: {n_pay} imported")

    if dry:
        print("[expense] DRY RUN — nothing written")
        db.rollback()
    else:
        db.commit()
        print("[expense] committed ✅")


# ================================================================ ehs

def migrate_ehs(dry: bool) -> None:
    from openpyxl import load_workbook

    from app.ehs.forms import ALL_FORMS
    from app.services import onedrive
    from app.services.ehs_excel_log import ehs_root

    db = SessionLocal()
    by_email, _ = _employee_maps(db)
    existing = {sid for (sid,) in db.query(EHSSubmission.submission_id).all()}
    total_new = 0

    for form in ALL_FORMS:
        path = f"{ehs_root()}/{form['folder']}/_MasterLog.xlsx"
        data = onedrive.download_from_path(path)
        if not data:
            continue
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Submissions"] if "Submissions" in wb.sheetnames else wb.active
        headers = [str(c.value or "") for c in ws[1]]

        def col(name):
            return headers.index(name) if name in headers else None

        n_fields = len(form["fields"])
        n_cl = len(form.get("checklist") or [])
        n_new = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            sid = str(row[0])
            if sid in existing:
                continue
            fields, photos = {}, {"fields": {}, "checklist": {}}
            for i, f in enumerate(form["fields"]):
                v = row[4 + i] if 4 + i < len(row) else None
                if f["type"] == "photo":
                    links = [u.strip() for u in str(v or "").split(",") if u.strip()]
                    if links:
                        photos["fields"][f["key"]] = [{"filename": None, "path": None, "webUrl": u} for u in links]
                else:
                    fields[f["key"]] = v if v is not None else ""
            checklist = []
            base = 4 + n_fields
            for i in range(n_cl):
                res = row[base + i * 3] if base + i * 3 < len(row) else ""
                rem = row[base + i * 3 + 1] if base + i * 3 + 1 < len(row) else ""
                link = row[base + i * 3 + 2] if base + i * 3 + 2 < len(row) else ""
                checklist.append({"result": res or "", "remarks": rem or ""})
                links = [u.strip() for u in str(link or "").split(",") if u.strip()]
                if links:
                    photos["checklist"][str(i)] = [{"filename": None, "path": None, "webUrl": u} for u in links]
            tail = 4 + n_fields + n_cl * 3
            pdf_link = row[tail] if tail < len(row) else None
            status = str(row[tail + 1] or "").strip().lower() if tail + 1 < len(row) else "approved"
            email = str(row[3] or "").strip().lower()
            emp = by_email.get(email)
            sub = EHSSubmission(
                submission_id=sid, form_id=form["id"], form_code=form["code"], form_title=form["title"],
                submitted_by_id=emp.id if emp else None,
                submitted_by_name=str(row[2] or "Unknown"), submitted_by_email=row[3],
                submitted_at_ist=str(row[1] or ""),
                fields=fields, checklist=checklist, photos=photos,
                status="approved" if status.startswith("appro") else "rejected",
                reviewed_by_name=row[tail + 2] if tail + 2 < len(row) else None,
                reviewed_by_email=row[tail + 3] if tail + 3 < len(row) else None,
                reviewed_at_ist=row[tail + 4] if tail + 4 < len(row) else None,
                edits_made=row[tail + 5] if tail + 5 < len(row) else None,
                reject_reason=row[tail + 6] if tail + 6 < len(row) else None,
                pdf_web_url=pdf_link,
            )
            if not dry:
                db.add(sub)
            existing.add(sid)
            n_new += 1
        if n_new:
            print(f"[ehs] {form['title']}: {n_new} imported from {path}")
        total_new += n_new

    print(f"[ehs] total: {total_new} historical submissions")
    if dry:
        print("[ehs] DRY RUN — nothing written")
        db.rollback()
    else:
        db.commit()
        print("[ehs] committed ✅")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("target", choices=["expense", "ehs"])
    ap.add_argument("sqlite", nargs="?", help="path to bsg-portal SQLite (expense only)")
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()
    if a.target == "expense":
        if not a.sqlite:
            sys.exit("expense migration needs the SQLite path")
        migrate_expense(a.sqlite, a.dry_run)
    else:
        migrate_ehs(a.dry_run)
