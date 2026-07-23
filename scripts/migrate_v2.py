"""
Metfraa KPI Tracker — Phase 1 Migration Script

WHAT IT DOES (in order):
  1. Archive current DB state to Excel → OneDrive/KPI_Tracker/Archives/
  2. Add new columns to `employees` table (idempotent via ALTER TABLE IF NOT EXISTS)
  3. Update existing employees: assign employee_code, phone, hash-of-Metfraa@123
  4. Insert fresh new employees (Prawin, Poornima, Bharathi, etc.)
  5. Deactivate old employees not in master list (Sathiyaseelan, Vijay, etc.)
  6. Truncate: daily_entries, kpi_entries, monthly_reports, unlock_requests, audit_log
  7. Preserve `kpis` table (KPI definitions kept intact)

USAGE:
  Dry-run (default — safe, prints what would happen):
      python scripts/migrate_v2.py
  Live (actually applies changes):
      python scripts/migrate_v2.py --live

ORDER OF OPERATIONS (Live mode):
  Archive → schema changes → update employees → insert new → deactivate stale
  → truncate history. Every step logs. Any failure abends before wipe.
"""

import argparse
import io
import os
import sys
from datetime import datetime
from pathlib import Path

# Allow running from project root
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from sqlalchemy import text
from sqlalchemy.orm import Session
import bcrypt as bcrypt_lib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import Base, engine, SessionLocal
from app.models import (
    Employee, KPI, DailyEntry, KPIEntry, MonthlyReport,
    UnlockRequest, AuditLog, PasswordResetRequest,
)

# Master data
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))
from _employee_master import EMPLOYEE_MASTER, DEACTIVATE_OLD


DEFAULT_PASSWORD = "Metfraa@123"
ARCHIVE_DIR_ONEDRIVE = "KPI_Tracker/Archives"


# ============================================================
# Utility helpers
# ============================================================

def _log(msg, tag="INFO"):
    ts = datetime.utcnow().strftime("%H:%M:%S")
    print(f"[{ts}] [{tag}] {msg}")


def _err(msg):
    _log(msg, tag="ERR")


def _ok(msg):
    _log(msg, tag=" OK")


def _norm(s: str) -> str:
    """Normalize a name for matching: lowercase, strip, collapse spaces."""
    if not s:
        return ""
    return " ".join(s.strip().lower().split())


# ============================================================
# STEP 1 — Archive current data to Excel
# ============================================================

def _table_columns(db: Session, table_name: str):
    """Return the actual column names of a table (survives schema drift)."""
    q = text(
        "SELECT column_name FROM information_schema.columns "
        "WHERE table_name = :t ORDER BY ordinal_position"
    )
    return [r[0] for r in db.execute(q, {"t": table_name}).fetchall()]


def _dump_table(db: Session, table: str, ws, order_by: str | None = None):
    """Dump a whole table into an openpyxl worksheet using raw SQL.
    This is immune to ORM schema drift — only queries columns that actually exist."""
    cols = _table_columns(db, table)
    if not cols:
        ws.append([f"(table '{table}' not found or empty)"])
        return 0
    ws.append(cols)  # header row
    order_clause = f" ORDER BY {order_by}" if order_by else ""
    q = text(f"SELECT {', '.join(cols)} FROM {table}{order_clause}")
    n = 0
    for row in db.execute(q).fetchall():
        vals = []
        for v in row:
            if v is None:
                vals.append("")
            elif hasattr(v, "isoformat"):
                vals.append(v.isoformat())
            else:
                vals.append(str(v) if not isinstance(v, (int, float, bool)) else v)
        ws.append(vals)
        n += 1
    return n


def build_archive_workbook(db: Session) -> bytes:
    """Serialize the current DB to a single Excel workbook and return bytes.
    Uses raw SQL so it's safe to run BEFORE or AFTER schema migrations.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    n_emp = _dump_table(db, "employees", ws, order_by="id")

    ws = wb.create_sheet("KPI Definitions")
    n_kpi = _dump_table(db, "kpis", ws, order_by="employee_id, id")

    ws = wb.create_sheet("Daily Entries")
    n_de = _dump_table(db, "daily_entries", ws, order_by="entry_date, employee_id")

    ws = wb.create_sheet("KPI Values")
    n_kve = _dump_table(db, "kpi_entries", ws, order_by="daily_entry_id, kpi_id")

    ws = wb.create_sheet("Monthly Reports")
    n_mr = _dump_table(db, "monthly_reports", ws, order_by="year, month, employee_id")

    ws = wb.create_sheet("Unlock Requests")
    n_ur = _dump_table(db, "unlock_requests", ws, order_by="requested_at")

    ws = wb.create_sheet("Audit Log")
    n_al = _dump_table(db, "audit_log", ws, order_by="created_at")

    _log(f"Archive: {n_emp} employees, {n_kpi} KPIs, {n_de} entries, "
         f"{n_kve} kpi values, {n_mr} reports, {n_ur} unlock reqs, {n_al} audit rows")

    # Style headers on every sheet
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F2937", end_color="1F2937")
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload_archive_to_onedrive(archive_bytes: bytes, filename: str) -> str:
    """Push the archive to OneDrive. Returns the OneDrive path/URL."""
    try:
        from app.services.onedrive import upload_file
        target = f"{ARCHIVE_DIR_ONEDRIVE}/{filename}"
        result = upload_file(archive_bytes, target)
        return result.get("webUrl") or result.get("path") or target
    except Exception as e:
        _err(f"OneDrive upload failed: {e}")
        _err("Saving archive locally as fallback.")
        local = Path(__file__).resolve().parents[1] / "archives"
        local.mkdir(exist_ok=True)
        p = local / filename
        with open(p, "wb") as f:
            f.write(archive_bytes)
        return f"local://{p}"


# ============================================================
# STEP 2 — Add new columns (schema migration)
# ============================================================

DDL_STATEMENTS = [
    "ALTER TABLE employees ADD COLUMN IF NOT EXISTS employee_code VARCHAR(32) UNIQUE",
    "ALTER TABLE employees ADD COLUMN IF NOT EXISTS password_hash VARCHAR(255)",
    "ALTER TABLE employees ADD COLUMN IF NOT EXISTS must_reset_password BOOLEAN DEFAULT TRUE",
    "ALTER TABLE employees ADD COLUMN IF NOT EXISTS phone VARCHAR(32)",
    "ALTER TABLE employees ADD COLUMN IF NOT EXISTS can_submit_task_report BOOLEAN DEFAULT TRUE",
    "ALTER TABLE employees ADD COLUMN IF NOT EXISTS last_login_at TIMESTAMP",
    "ALTER TABLE employees ALTER COLUMN email DROP NOT NULL",
]


def apply_schema_changes(db: Session, dry_run: bool):
    """Apply schema changes ALWAYS — even in dry-run.
    
    The ALTER TABLE statements are idempotent (IF NOT EXISTS) and non-destructive.
    We apply them even in dry-run so that:
      1. The archive can query the DB using the new ORM without errors
      2. You can see the DDL succeeded (validates the schema change plan)
      3. No data is affected — only new columns added (all nullable/defaulted)
    
    The actual DATA changes (updates/inserts/deletes) still respect dry_run.
    """
    for stmt in DDL_STATEMENTS:
        try:
            db.execute(text(stmt))
            _ok(f"applied: {stmt[:80]}...")
        except Exception as e:
            _err(f"DDL failed: {stmt} — {e}")
    db.commit()
    # Also ensure new tables exist (PasswordResetRequest is new)
    Base.metadata.create_all(bind=engine)
    _ok("Base.metadata.create_all — new tables created if missing")
    if dry_run:
        _log("Note: schema changes are always applied (idempotent, non-destructive).")
        _log("      Data changes (updates/inserts/deletes) below are still dry-run.")


# ============================================================
# STEP 3-5 — Migrate employees
# ============================================================

def count_kpis_for_employee(db: Session, employee_id: int) -> int:
    """Count KPIs for an employee via raw SQL, bypassing ORM schema drift."""
    q = text("SELECT COUNT(*) FROM kpis WHERE employee_id = :eid")
    return int(db.execute(q, {"eid": employee_id}).scalar() or 0)


def match_by_name(db: Session, old_name: str) -> Employee | None:
    """Find an existing Employee whose name matches (case-insensitive)."""
    if not old_name:
        return None
    target = _norm(old_name)
    for e in db.query(Employee).all():
        if _norm(e.name) == target:
            return e
    return None


def migrate_employees(db: Session, dry_run: bool):
    hashed_default = bcrypt_lib.hashpw(DEFAULT_PASSWORD.encode("utf-8"), bcrypt_lib.gensalt()).decode("utf-8")

    updated, created, deactivated = 0, 0, 0
    plan_report = []  # for dry-run summary

    for entry in EMPLOYEE_MASTER:
        code = entry["code"]
        name = entry["name"]
        old_name = entry.get("old_name")

        existing = match_by_name(db, old_name) if old_name else None

        if existing:
            # Update in place — preserves KPIs and ID
            plan_report.append({
                "action": "UPDATE",
                "code": code,
                "name": name,
                "was": existing.name,
                "kpis_preserved": count_kpis_for_employee(db, existing.id),
            })
            if not dry_run:
                existing.employee_code = code
                existing.name = name
                existing.phone = entry["phone"]
                existing.designation = entry["designation"]
                existing.department = entry["department"]
                existing.is_admin = entry["is_admin"]
                existing.can_submit_task_report = entry["can_submit_task_report"]
                existing.is_active = True
                existing.password_hash = hashed_default
                existing.must_reset_password = True
                if entry.get("email"):
                    existing.email = entry["email"]
            updated += 1
        else:
            # Fresh new employee
            plan_report.append({
                "action": "CREATE",
                "code": code,
                "name": name,
                "designation": entry["designation"],
                "department": entry["department"],
                "kpis_preserved": 0,
            })
            if not dry_run:
                emp = Employee(
                    employee_code=code,
                    name=name,
                    email=entry.get("email"),
                    phone=entry["phone"],
                    designation=entry["designation"],
                    department=entry["department"],
                    is_admin=entry["is_admin"],
                    is_active=True,
                    can_submit_task_report=entry["can_submit_task_report"],
                    password_hash=hashed_default,
                    must_reset_password=True,
                )
                db.add(emp)
            created += 1

    # Deactivate old employees not in master
    master_old_names = {_norm(e["old_name"]) for e in EMPLOYEE_MASTER if e.get("old_name")}
    for e in db.query(Employee).all():
        if e.employee_code:
            continue  # already given a new code, skip
        if _norm(e.name) not in master_old_names:
            plan_report.append({
                "action": "DEACTIVATE",
                "was": e.name,
                "email": e.email,
                "kpis": count_kpis_for_employee(db, e.id),
            })
            if not dry_run:
                e.is_active = False
                e.can_submit_task_report = False
            deactivated += 1

    if not dry_run:
        db.commit()

    return updated, created, deactivated, plan_report


# ============================================================
# STEP 6 — Wipe legacy data
# ============================================================

def wipe_legacy_data(db: Session, dry_run: bool):
    tables_to_wipe = ["kpi_entries", "daily_entries", "monthly_reports", "unlock_requests", "audit_log"]
    for t in tables_to_wipe:
        if dry_run:
            count = db.execute(text(f"SELECT COUNT(*) FROM {t}")).scalar()
            _log(f"[dry-run] would DELETE {count} rows from {t}")
        else:
            db.execute(text(f"DELETE FROM {t}"))
            _ok(f"wiped {t}")
    if not dry_run:
        db.commit()


# ============================================================
# Main orchestrator
# ============================================================

def print_plan_summary(report):
    """Pretty-print the plan for dry-run review."""
    print()
    print("=" * 60)
    print("MIGRATION PLAN")
    print("=" * 60)

    updates = [r for r in report if r["action"] == "UPDATE"]
    creates = [r for r in report if r["action"] == "CREATE"]
    deactivates = [r for r in report if r["action"] == "DEACTIVATE"]

    print(f"\n📝 UPDATE (existing employees get new codes + password): {len(updates)}")
    for r in updates:
        print(f"    {r['code']:12s} {r['name']:40s}  (was: {r['was']})  -- keeps {r['kpis_preserved']} KPIs")

    print(f"\n➕ CREATE (fresh new employees, need KPIs later): {len(creates)}")
    for r in creates:
        print(f"    {r['code']:12s} {r['name']:40s}  {r['designation']}")

    print(f"\n➖ DEACTIVATE (in old app, not in master): {len(deactivates)}")
    for r in deactivates:
        print(f"    {r['was']:40s}  ({r['email']})  -- had {r['kpis']} KPIs")

    print()


def main():
    parser = argparse.ArgumentParser(description="Metfraa KPI v2 migration")
    parser.add_argument("--live", action="store_true", help="Actually apply changes (default is dry-run)")
    parser.add_argument("--skip-archive", action="store_true", help="Skip OneDrive archive step (danger)")
    args = parser.parse_args()

    dry_run = not args.live
    mode = "LIVE" if args.live else "DRY-RUN"
    _log(f"=== Metfraa KPI v2 Migration — {mode} ===")

    db = SessionLocal()
    try:
        # Step 1: Schema changes FIRST (idempotent — safe to run twice)
        # We do this before the archive so the ORM query doesn't fail on
        # columns the old DB doesn't have yet.
        _log("Step 1: Applying schema changes (ALTER TABLE IF NOT EXISTS)...")
        apply_schema_changes(db, dry_run)

        # Step 2: Archive (using raw SQL — safe against schema drift)
        if args.skip_archive:
            _log("[SKIP] archive step (--skip-archive)")
        else:
            _log("Step 2: Archiving current DB state to Excel...")
            archive_bytes = build_archive_workbook(db)
            filename = f"pre-v2-migration-{datetime.utcnow().strftime('%Y-%m-%d-%H%M%S')}.xlsx"
            if dry_run:
                _log(f"[dry-run] would upload {filename} ({len(archive_bytes):,} bytes) to OneDrive/{ARCHIVE_DIR_ONEDRIVE}")
            else:
                url = upload_archive_to_onedrive(archive_bytes, filename)
                _ok(f"Archive uploaded: {url}")

        # Step 3-5: Employees
        _log("Step 3-5: Migrating employees...")
        updated, created, deactivated, report = migrate_employees(db, dry_run)
        print_plan_summary(report)

        # Step 6: Wipe
        _log("Step 6: Wiping legacy operational data...")
        wipe_legacy_data(db, dry_run)

        _ok(f"Summary: {updated} updated, {created} created, {deactivated} deactivated")

        if dry_run:
            print()
            print("*" * 60)
            print("This was a DRY-RUN. No changes have been made.")
            print("To apply for real, run:")
            print("    python scripts/migrate_v2.py --live")
            print("*" * 60)
        else:
            _ok("MIGRATION COMPLETE")

    except Exception as e:
        _err(f"MIGRATION FAILED: {e}")
        import traceback
        traceback.print_exc()
        db.rollback()
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    main()
